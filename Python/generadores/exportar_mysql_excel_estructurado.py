"""
referencia_conexion_excel.py
Plantilla de consulta y exportación a Excel con tabla estructurada.
"""

# 📦 Librerías necesarias
import mysql.connector
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# 🔌 1. Conexión a MySQL
conexion = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="ventas_integrado"
)

# 🧾 2. Consulta con JOIN = Une la tabla ventas con clientes y productos
#Unir varias tablas relacionales (ventas + clientes + productos) 
#en un solo resultado final listo para análisis o visualización.
#Relaciona por las llaves foráneas: id_cliente y id_producto
# v.id_venta, v.fecha, v.cantidad, v.precio_unitario: son del hecho de la venta
# c.nombre, c.ciudad: vienen de la dimensión clientes
# p.nombre, p.categoria: vienen de la dimensión productos
# v.cantidad * v.precio_unitario AS total:_venta agrega la columna calculada del total vendido por fila
consulta = """
SELECT
  v.id_ventas,
  c.nombre AS cliente,
  c.zona,
  p.nombre AS producto,
  p.categoria,
  v.fecha,
  v.cantidad,
  v.precio_unitario,
  (v.cantidad * v.precio_unitario) AS total_venta
FROM ventas v
JOIN clientes c ON v.id_cliente = c.id_cliente
JOIN productos p ON v.id_producto = p.id_producto
"""

# 🧮 3. Ejecutar la consulta y convertir a DataFrame
df = pd.read_sql(consulta, conexion)
print(df.head())  # Muestra para verificación

# 4️⃣ Crear nombre dinámico con fecha
fecha_hoy = datetime.now().strftime('%Y_%m_%d')
nombre_archivo = f"entregas_{fecha_hoy}.xlsx"

# 📤 5. Exportar a Excel
archivo_excel = "ventas_limpias.xlsx"
df.to_excel(archivo_excel, index=False)

# ✅ 6. Crear tabla estructurada automáticamente (versión optimizada)
def convertir_a_tabla_excel(ruta_archivo, nombre_tabla):
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    col_final = chr(64 + max_col)  # Calcula la última columna (A, B, C...)

    rango = f"A1:{col_final}{max_row}"  # Define el rango A1:col_final
    tabla = Table(displayName=nombre_tabla, ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    wb.save(ruta_archivo)
    print(f"✅ Tabla '{nombre_tabla}' creada en {ruta_archivo}")

# Llamar la función
convertir_a_tabla_excel(archivo_excel, "VentasTabla")

# 🔒 6. Cerrar la conexión
conexion.close()

# ¿Qué hace este script?
# 🔄 Funciona como un puente entre la base de datos (MySQL) y Power BI.
#   🔎 Lee los datos desde MySQL
#   Une las tres tablas (ventas, clientes, productos)
#   Calcula el campo total (cantidad * precio_unitario)
#  📤 Exporta esos datos a un Excel
#  Lo guarda como ventas_exportadas.xlsx
#  📈 Luego Power BI lee ese Excel
#  Y tú ya puedes hacer:
#    Tarjetas
#    Gráficos por ciudad, categoría, año...
#    Medidas DAX como TotalVentas, Ventas LY, etc.

# Este código genera un archivo Excel limpio y automatizado para que Power BI 
# pueda leerlo como si viniera de SAP, SQL Server o cualquier fuente empresarial.