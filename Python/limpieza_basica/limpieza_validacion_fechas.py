"""
referencia_limpieza_con_fecha.py
===============================
🧹 Script de referencia para limpiar datos y agregar columna de fecha

✔ Limpieza básica con pandas
✔ Generación de columna 'fecha' ideal para Power BI (ventas por mes, año, acumulado)
✔ Exporta archivos como tablas estructuradas en Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# 📦 Función: convertir archivo Excel a tabla estructurada
def convertir_a_tabla_excel(ruta_archivo, nombre_tabla):
    wb = load_workbook(ruta_archivo)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    col_final = get_column_letter(max_col)
    rango = f"A1:{col_final}{max_row}"
    tabla = Table(displayName=nombre_tabla, ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)
    wb.save(ruta_archivo)
    print(f"✅ Tabla '{nombre_tabla}' creada en {ruta_archivo}")

# 📥 Cargar archivo original con varias hojas
archivo = 'datos_ejemplo.xlsx'
ventas = pd.read_excel(archivo, sheet_name='ventas')
clientes = pd.read_excel(archivo, sheet_name='clientes')
productos = pd.read_excel(archivo, sheet_name='productos')

# 🧼 Limpieza de datos
ventas.dropna(inplace=True)
clientes.dropna(inplace=True)
productos.dropna(inplace=True)

clientes["nombre_cliente"] = clientes["nombre_cliente"].str.strip()
productos["categoria"] = productos["categoria"].str.strip()

ventas["total"] = ventas["cantidad"] * ventas["precio_unitario"]

# 📆 Agregar columna de fecha simulada dinámica
# Ideal para relacionar con una tabla calendario en Power BI
hoy = datetime.today()
fecha_base = datetime(hoy.year, hoy.month, 1)
ventas['fecha'] = [fecha_base + timedelta(days=i) for i in range(len(ventas))]

# 💾 Exportar archivos
ventas_path = 'ventas_limpias_20250615.xlsx'
clientes_path = 'clientes_limpios_20250615.xlsx'
productos_path = 'productos_limpios_20250615.xlsx'

ventas.to_excel(ventas_path, index=False)
clientes.to_excel(clientes_path, index=False)
productos.to_excel(productos_path, index=False)

# 📋 Convertir a tabla estructurada (Power BI friendly)
convertir_a_tabla_excel(ventas_path, "Ventas20250615")
convertir_a_tabla_excel(clientes_path, "Clientes20250615")
convertir_a_tabla_excel(productos_path, "Productos20250615")

print("🎯 Limpieza completada y tablas listas para Power BI.")

