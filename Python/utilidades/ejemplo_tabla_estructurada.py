"""
procesar_consolidado_excel.py
Limpieza y conversión a tabla estructurada de tres hojas: ventas, clientes, productos
"""
"""
ejemplo_tabla_estructurada.py
⚙️ Plantilla práctica para limpiar y convertir a tabla Excel rápidamente.
Ideal para pruebas locales, ejercicios técnicos o flujo básico sin validación.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# 📌 Función para convertir a tabla estructurada
#def: Es una palabra clave de Python que significa “define una función”
#convertir_a_tabla_excel:Es el nombre que le damos a la función
#(ruta_archivo, nombre_tabla):	Son los parámetros que la función necesita para funcionar

def convertir_a_tabla_excel(ruta_archivo, nombre_tabla):
    # Cargar el archivo Excel existente
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    # Obtener última fila y columna con datos
    max_row = ws.max_row
    max_col = ws.max_column

    # Convertir número de columna a letra (A, B, ..., Z, AA, AB, etc.)
    col_final = get_column_letter(max_col)

    # Definir el rango de la tabla (ej: A1:G11)
    rango = f"A1:{col_final}{max_row}"

    # Crear la tabla con estilo visual
    tabla = Table(displayName=nombre_tabla, ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    # Guardar cambios
    wb.save(ruta_archivo)
    print(f"✅ Tabla '{nombre_tabla}' creada en {ruta_archivo}")

# 📥 1. Cargar hojas del archivo original
archivo_original = "consolidado.xlsx"
ventas = pd.read_excel(archivo_original, sheet_name='ventas')
clientes = pd.read_excel(archivo_original, sheet_name='clientes')
productos = pd.read_excel(archivo_original, sheet_name='productos')

# 🧹 2. Limpieza básica

# Ventas
ventas.dropna(inplace=True)
ventas['total'] = ventas['cantidad'] * ventas['precio_unitario']

# Clientes
clientes.dropna(inplace=True)
clientes['nombre_cliente'] = clientes['nombre_cliente'].str.strip()
clientes['ciudad'] = clientes['ciudad'].str.strip()

# Productos
productos.dropna(inplace=True)
productos['nombre_producto'] = productos['nombre_producto'].str.strip()
productos['categoria'] = productos['categoria'].str.strip()

# 💾 3. Guardar cada hoja en su archivo limpio
ventas_path = "ventas_limpias.xlsx"
clientes_path = "clientes_limpios.xlsx"
productos_path = "productos_limpios.xlsx"

ventas.to_excel(ventas_path, index=False)
clientes.to_excel(clientes_path, index=False)
productos.to_excel(productos_path, index=False)

# 📋 4. Convertir cada archivo a tabla estructurada
convertir_a_tabla_excel(ventas_path, "Ventas")
convertir_a_tabla_excel(clientes_path, "Clientes")
convertir_a_tabla_excel(productos_path, "Productos")

# 🎉 Confirmación final
print("✅ Todo listo: archivos limpios y estructurados creados correctamente.")
