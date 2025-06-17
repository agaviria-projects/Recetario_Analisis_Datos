"""
plantilla_procesar_consolidado.py
Versión mejorada para limpiar archivos Excel con múltiples hojas (ventas, clientes, productos)
y convertir cada hoja en un archivo individual estructurado como tabla Excel.
"""

from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# 🧰 Función para convertir un archivo Excel en tabla estructurada
def convertir_a_tabla_excel(ruta_archivo, nombre_tabla):
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    col_final = chr(64 + max_col)  # A = 65

    rango = f"A1:{col_final}{max_row}"
    tabla = Table(displayName=nombre_tabla, ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    wb.save(ruta_archivo)
    print(f"✅ Tabla '{nombre_tabla}' creada en {ruta_archivo}")

# 🧹 Función para limpiar columnas (quita espacios adelante y atrás)
def limpiar_columnas(df, columnas):
    for col in columnas:
        df[col] = df[col].str.strip()

# 📌 1. CAMBIA ESTE NOMBRE POR EL ARCHIVO QUE TE ENTREGUEN
# Ejemplo: archivo_original = "ventas_junio.xlsx"
archivo_original = "consolidado.xlsx"

# 📆 Fecha actual para incluir en los nombres de archivos de salida
fecha = datetime.today().strftime("%Y%m%d")

# 🔍 2. VALIDAR QUE EL ARCHIVO EXISTA
if not os.path.exists(archivo_original):
    print(f"❌ El archivo {archivo_original} no fue encontrado.")
    exit()

# ---------------------------------------
# 📥 3. CARGA Y LIMPIEZA DE DATOS
# CAMBIA LOS NOMBRES DE HOJAS SEGÚN TU ARCHIVO
# ---------------------------------------

ventas = pd.read_excel(archivo_original, sheet_name='ventas')
clientes = pd.read_excel(archivo_original, sheet_name='clientes')
productos = pd.read_excel(archivo_original, sheet_name='productos')

# 🔧 LIMPIEZA
ventas.dropna(inplace=True)
ventas['total'] = ventas['cantidad'] * ventas['precio_unitario']

clientes.dropna(inplace=True)
limpiar_columnas(clientes, ["nombre_cliente", "ciudad"])

productos.dropna(inplace=True)
limpiar_columnas(productos, ["nombre_producto", "categoria"])

# ---------------------------------------
# 💾 4. GUARDAR ARCHIVOS LIMPIOS CON FECHA
# Puedes cambiar los nombres si lo deseas
# ---------------------------------------

ventas_path = f"ventas_limpias_{fecha}.xlsx"
clientes_path = f"clientes_limpios_{fecha}.xlsx"
productos_path = f"productos_limpios_{fecha}.xlsx"

ventas.to_excel(ventas_path, index=False)
clientes.to_excel(clientes_path, index=False)
productos.to_excel(productos_path, index=False)

# ---------------------------------------
# 📋 5. CONVERTIR A TABLAS ESTRUCTURADAS
# Puedes cambiar el nombre de la tabla final (para Power BI)
# ---------------------------------------

convertir_a_tabla_excel(ventas_path, f"Ventas{fecha}")
convertir_a_tabla_excel(clientes_path, f"Clientes{fecha}")
convertir_a_tabla_excel(productos_path, f"Productos{fecha}")

# ✅ Confirmación final
print("🎉 Archivos limpios y estructurados creados correctamente.")
